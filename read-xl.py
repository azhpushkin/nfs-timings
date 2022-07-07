from datetime import timedelta

import os
import django
from openpyxl import load_workbook

# Setup django to use models
os.environ['DJANGO_SETTINGS_MODULE'] = 'timings.settings'
django.setup()

from django.utils import timezone

from stats.models import BoardRequest
from stats.processing import time_to_int, process_lap_lime



wb = load_workbook(filename='simulation/race.xlsx')
print('loaded!')

data = []
q = 0

teamnumber = 0

for sheet in wb.worksheets[2:]:
    # iterate over teams
    cur_col = 3
    cur_time = 0
    teamnumber += 1

    while sheet.cell(row=20, column=cur_col).value:
        # iterate over stints
        ontrack = 0

        teamname = sheet['A4'].value
        print(teamname)

        name_link = sheet.cell(row=20, column=cur_col).value[1:]
        name_cell = sheet[name_link]
        name = name_cell.value

        # SKip karts like `4,44` which indicate something failing
        try:
            kart = int(name_cell.offset(column=1).value)
        except:
            cur_col += 1
            continue

        if cur_col != 3:
            new_time = time_to_int(str(name_cell.offset(column=7, row=-1).value))
            cur_time = new_time

        cur_row = 21
        while sheet.cell(row=cur_row, column=cur_col).value:
            # iterate over laps
            time = float(sheet.cell(row=cur_row, column=cur_col).value)

            dummy_board_request = BoardRequest.objects.create(
                created_at=timezone.now() + timedelta(seconds=int(cur_time)),
                url='http://example.com',
                status=200,
                response='lala',
                response_json={},
                is_processed=True
            )

            process_lap_lime(
                dummy_board_request,
                race_time=cur_time,
                team_number=teamnumber,
                team_name=teamname,
                pilot_name=name,
                kart=kart,
                ontrack=ontrack,
                lap_time=time
            )
            cur_row += 1
            cur_time = round(cur_time + time, 3)
            ontrack = round(ontrack + time, 3)
            q += 1
        cur_col+=1

print("Total", q)

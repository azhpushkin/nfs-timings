import json
from openpyxl import load_workbook

wb = load_workbook(filename='race.xlsx')
print('loaded!')


def time_to_int(t):
    h, m, s = t.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)


def int_to_time(t):
    t = int(t)
    h = t // 3600
    m = (t - 3600 * h) // 60
    s = t - (h*3600) - (m * 60)
    return f'{h}:{m}:{s}'


data = []
q = 0

for sheet in wb.worksheets[2:]:
    cur_col = 3
    cur_time = 0
    while sheet.cell(row=20, column=cur_col).value:

        name_link = sheet.cell(row=20, column=cur_col).value[1:]
        name_cell = sheet[name_link]
        name = name_cell.value

        # SKip karts like `4,44` which indicate something failing
        try:
            kart = int(name_cell.offset(column=1).value)
        except:
            cur_col += 1
            continue

        if cur_col != 3:
            new_time = time_to_int(str(name_cell.offset(column=7, row=-1).value))
            cur_time = new_time

        cur_row = 21
        while sheet.cell(row=cur_row, column=cur_col).value:
            time = float(sheet.cell(row=cur_row, column=cur_col).value)
            # print(f'{name} on kart {kart} (at {int_to_time(cur_time)}), {time}')
            # input()
            cur_row += 1
            cur_time = round(cur_time + time, 3)
            q += 1
        cur_col+=1

print("Total", q)
